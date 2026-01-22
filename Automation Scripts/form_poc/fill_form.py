import re
import sys
from pathlib import Path
from typing import Dict, Any

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# Update these
FORM_URL = "https://docs.google.com/forms/d/e/1FAIpQLScTWvccbONPrWjysaRVlR0DdqljxnJlWTrn83kqxoSt0uMNtg/viewform"
EXCEL_PATH = Path("./data.xlsx")

# Map Excel headers -> form field selector
# Replace selectors once you capture them (e.g. from Playwright codegen)
SELECTORS = {
    "Hotel Code": "role=textbox[name='Hotel Code']",
    "PM": "role=textbox[name='PM']",
    "PM Phone Number": "role=textbox[name='PM Phone Number']",
    "Primary Admin Name": "role=textbox[name='Primary Admin Name']",
    "Primary Admin UN": "role=textbox[name='Primary Admin UN']",
    "Primary Admin Email": "role=textbox[name='Primary Admin Email']",
}

# Adjust regex to match the on-page order number text
ORDER_REGEX = re.compile(r"\b(\d{6,})\b")


def read_row_by_order_number(xlsx_path: Path, order_number: str) -> Dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    normalized_headers = [
        "" if h is None else str(h).strip().lower() for h in headers
    ]
    if "order number" not in normalized_headers:
        raise ValueError("Missing 'Order number' header in Excel.")

    order_idx = normalized_headers.index("order number")

    def normalize_order(value) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    target = normalize_order(order_number)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[order_idx] is None:
            continue
        if normalize_order(row[order_idx]) == target:
            return {headers[i]: row[i] for i in range(len(headers))}

    raise ValueError(f"Order number {order_number} not found in Excel.")


def extract_order_number(text: str) -> str:
    match = ORDER_REGEX.search(text or "")
    if not match:
        raise ValueError("Could not find order number on the page. Update ORDER_REGEX.")
    return match.group(1)


def main():
    xlsx_path = EXCEL_PATH
    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1])

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        page.goto(FORM_URL, wait_until="domcontentloaded")

        # Give the page a moment to render its visible text
        page.wait_for_timeout(1000)
        page_text = page.inner_text("body")
        order_number = extract_order_number(page_text)

        row = read_row_by_order_number(xlsx_path, order_number)

        # Fill fields
        for header, selector in SELECTORS.items():
            if header not in row:
                continue
            value = "" if row[header] is None else str(row[header])
            if not selector:
                continue
            page.fill(selector, value)

        # Example: submit (update selector)
        # page.click("#submit")

        # Keep browser open for inspection
        print("Check the browser, then press Enter to close...")
        input()
        browser.close()



if __name__ == "__main__":
    main()
